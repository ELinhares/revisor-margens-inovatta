import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

REQUIRED_COLUMNS = ["Código do Produto", "Produto", "Venda (R$)", "Margem Atual"]

COLUMN_ALIASES = {
    "Produto (descrição)": "Produto",
    "Produto (Descrição)": "Produto",
    "produto": "Produto",
    "Descrição": "Produto",
    "Descricao": "Produto",
    "descricao": "Produto",
    "Descrição do Produto": "Produto",
    "Nome do Produto": "Produto",
    "codigo": "Código do Produto",
    "Codigo": "Código do Produto",
    "Codigo do Produto": "Código do Produto",
    "Cod. Produto": "Código do Produto",
    "Cód. Produto": "Código do Produto",
    "Código": "Código do Produto",
    "COD": "Código do Produto",
    "SKU": "Código do Produto",
    "Ref": "Código do Produto",
    "Referência": "Código do Produto",
    "venda": "Venda (R$)",
    "Venda": "Venda (R$)",
    "Vendas": "Venda (R$)",
    "Receita": "Venda (R$)",
    "Faturamento": "Venda (R$)",
    "Fat.": "Venda (R$)",
    "Valor Vendido": "Venda (R$)",
    "Total Vendido": "Venda (R$)",
    "margem": "Margem Atual",
    "Margem": "Margem Atual",
    "Margem (%)": "Margem Atual",
    "Margem%": "Margem Atual",
    "MG": "Margem Atual",
    "Mg.": "Margem Atual",
    "Margem de Contribuição": "Margem Atual",
}


def _fuzzy_match(col: str) -> str | None:
    c = col.lower().strip()
    if any(k in c for k in ["cód", "cod", "código", "codigo", "ref", "sku"]):
        return "Código do Produto"
    if any(k in c for k in ["prod", "desc", "nome", "item", "mercad"]):
        return "Produto"
    if any(k in c for k in ["vend", "receita", "fatur", "valor", "total"]):
        return "Venda (R$)"
    if any(k in c for k in ["margem", "margin", " mg", "lucro", "contrib"]):
        return "Margem Atual"
    return None


def _infer_mapping(columns: list[str]) -> dict[str, str | None]:
    result: dict[str, str | None] = {}
    assigned: set[str] = set()
    for col in columns:
        if col in REQUIRED_COLUMNS:
            result[col] = col
            assigned.add(col)
        elif col in COLUMN_ALIASES:
            target = COLUMN_ALIASES[col]
            if target not in assigned:
                result[col] = target
                assigned.add(target)
            else:
                result[col] = None
        else:
            target = _fuzzy_match(col)
            if target and target not in assigned:
                result[col] = target
                assigned.add(target)
            else:
                result[col] = None
    return result


def _strip_percent(series: pd.Series) -> pd.Series:
    """Remove % character and strip whitespace from string values in a series."""
    if series.dtype == object:
        return series.astype(str).str.replace("%", "", regex=False).str.strip()
    return series


def _detect_margin_format(series: pd.Series) -> str:
    """Return 'decimal' if values look like 0.XX form, 'percent' otherwise.

    Heuristic: if the 75th percentile of absolute non-zero values is < 1.5,
    the column is almost certainly in decimal form (e.g. 0.185 = 18.5%).
    """
    valid = pd.to_numeric(_strip_percent(series), errors="coerce").dropna()
    valid = valid[valid != 0]
    if len(valid) == 0:
        return "percent"
    p75 = valid.abs().quantile(0.75)
    return "decimal" if p75 < 1.5 else "percent"


def check_columns(file_bytes: bytes) -> dict:
    df = pd.read_excel(BytesIO(file_bytes), engine="openpyxl")
    all_columns = [str(c).strip() for c in df.columns]
    df.columns = all_columns

    inferred = _infer_mapping(all_columns)
    mapped_required = {v for v in inferred.values() if v}

    found = [c for c in REQUIRED_COLUMNS if c in mapped_required]
    missing = [c for c in REQUIRED_COLUMNS if c not in mapped_required]

    # Detect margin format using inferred margem column
    margin_col_in_file = next(
        (k for k, v in inferred.items() if v == "Margem Atual"), None
    )
    if margin_col_in_file and margin_col_in_file in df.columns:
        margin_format = _detect_margin_format(df[margin_col_in_file])
    else:
        margin_format = "percent"

    df_preview = df.rename(columns={k: v for k, v in inferred.items() if v})
    preview_cols = [c for c in REQUIRED_COLUMNS if c in df_preview.columns]
    preview = df_preview[preview_cols].head(5).fillna("").to_dict(orient="records")

    return {
        "all_columns": all_columns,
        "inferred_mapping": inferred,
        "columns_found": found,
        "missing_columns": missing,
        "margin_format": margin_format,
        "total_rows": len(df),
        "preview": preview,
    }


def validate_and_read(
    file_bytes: bytes,
    column_mapping: dict[str, str] | None = None,
    margin_format: str | None = None,
) -> pd.DataFrame:
    df = pd.read_excel(BytesIO(file_bytes), engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]

    if column_mapping:
        df.rename(columns={k: v for k, v in column_mapping.items() if v}, inplace=True)
    else:
        df.rename(columns=COLUMN_ALIASES, inplace=True)
        still_missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
        if still_missing:
            inferred = _infer_mapping(list(df.columns))
            fallback = {k: v for k, v in inferred.items() if v in still_missing}
            df.rename(columns=fallback, inplace=True)

    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"Colunas obrigatórias ausentes: {missing}")

    df["Venda (R$)"] = pd.to_numeric(df["Venda (R$)"], errors="coerce").fillna(0.0)
    df["Margem Atual"] = pd.to_numeric(_strip_percent(df["Margem Atual"]), errors="coerce").fillna(0.0)

    # Use the format detected during /validate if provided; otherwise re-detect
    fmt = margin_format if margin_format in ("decimal", "percent") else _detect_margin_format(df["Margem Atual"])
    if fmt == "decimal":
        df["Margem Atual"] = df["Margem Atual"] * 100

    return df.copy()


def write_excel(df: pd.DataFrame) -> bytes:
    output = BytesIO()
    col_order = ["Código do Produto", "Produto", "Venda (R$)", "Margem Atual", "ABC", "Margem Sugerida"]
    extra = [c for c in df.columns if c not in col_order]
    final_cols = [c for c in col_order if c in df.columns] + extra
    df = df[final_cols]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Revisão de Margens")
        ws = writer.sheets["Revisão de Margens"]

        header_fill = PatternFill(start_color="0057B8", end_color="0057B8", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col_idx, col_name in enumerate(final_cols, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                col_name = final_cols[cell.column - 1] if cell.column <= len(final_cols) else ""
                if col_name == "Venda (R$)":
                    cell.number_format = 'R$ #,##0.00'
                elif col_name in ("Margem Atual", "Margem Sugerida"):
                    cell.number_format = '0.00"%"'

        for col_idx, col_name in enumerate(final_cols, start=1):
            max_len = max(
                len(str(col_name)),
                *[len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row + 1)],
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        ws.freeze_panes = "A2"

    return output.getvalue()
